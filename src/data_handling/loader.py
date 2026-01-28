"""
Handles loading and preprocessing of input data from CSV or Excel files.

This module is responsible for:
- Reading data using pandas, with a "smart read" capability to handle files
  with an unknown number of trailing empty rows.
- Standardizing column names based on configurable input profiles.
- Initializing new columns required by the pipeline (e.g., for status tracking,
  extracted data, and run identifiers).
- Performing initial normalization of phone numbers if a designated phone
  number column is present in the input.
"""
import csv # For smart CSV reading
import logging
import uuid # For RunID
from typing import Optional, List, Dict, Any, Union, Iterable

import pandas as pd
from openpyxl import load_workbook # For smart Excel reading

# Import AppConfig directly. Its __init__ handles .env loading.
# If this import fails, it's a critical setup error for the application.
from ..core.config import AppConfig

# Configure logging.
# The setup_logging() function might rely on environment variables that are
# loaded when AppConfig is instantiated.
try:
    from ..core.logging_config import setup_logging
    # AppConfig() is instantiated globally in config.py if needed by other modules,
    # or when an instance is created. Here, we just ensure logging is set up.
    setup_logging()
    logger = logging.getLogger(__name__)
except ImportError:
    # Fallback basic logging configuration if core.logging_config is unavailable.
    # This might happen during isolated testing or if there's a setup issue.
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    logger = logging.getLogger(__name__)
    logger.warning( # Changed from info to warning as this is a fallback
        "Basic logging configured for loader.py due to missing "
        "core.logging_config or its dependencies. This is a fallback."
    )


def _coalesce_duplicate_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Ensure df has unique column names.

    Some workflows intentionally feed "already augmented" CSVs back into the pipeline.
    When an input profile maps (e.g.) "Website" -> "GivenURL" but the input already
    contains a "GivenURL" column, pandas ends up with duplicate column names.

    Duplicate column names are dangerous because row.get("GivenURL") can return a
    pandas Series (multiple values), which then breaks downstream truthiness checks.

    Strategy:
    - For each duplicated column name, coalesce row-wise by taking the first non-empty value
      across the duplicate columns.
    - Drop all duplicates and re-insert a single coalesced column at the position of the
      first occurrence (to preserve roughly stable column ordering).
    """
    try:
        cols = list(df.columns)
        dup_names = [c for c in cols if cols.count(c) > 1]
        if not dup_names:
            return df
        seen = set()
        for name in dup_names:
            if name in seen:
                continue
            seen.add(name)
            # Capture the first occurrence index before we drop anything
            try:
                first_idx = cols.index(name)
            except ValueError:
                first_idx = len(df.columns)
            sub = df.loc[:, df.columns == name]
            if sub.shape[1] <= 1:
                continue
            # Treat empty/whitespace-only strings as missing for coalescing
            try:
                sub_clean = sub.replace(r"^\s*$", pd.NA, regex=True)
            except Exception:
                sub_clean = sub
            coalesced = sub_clean.bfill(axis=1).iloc[:, 0]
            # Keep a stable empty-string for missing values (we run with keep_default_na=False)
            try:
                coalesced = coalesced.fillna("")
            except Exception:
                pass
            # Drop all columns with this label and re-insert one coalesced column
            df = df.drop(columns=[name])
            # Bound insertion index to current width
            insert_at = min(first_idx, len(df.columns))
            df.insert(insert_at, name, coalesced)
            cols = list(df.columns)
        return df
    except Exception:
        # Best-effort only; never fail the loader due to dedupe logic
        return df


def _is_row_empty(row_values: Iterable[Any]) -> bool:
    """
    Checks if all values in a given row are effectively empty.

    An empty value is defined as None, an empty string, or a string
    containing only whitespace.

    Args:
        row_values: An iterable (e.g., list, tuple) of values in a row.

    Returns:
        True if all values in the row are empty, False otherwise.
        Returns True if row_values itself is None or an empty iterable.
    """
    if not row_values: # Handles case where row_values might be None or an empty list/tuple
        return True
    return all(pd.isna(value) or (isinstance(value, str) and not value.strip()) for value in row_values)


def _detect_csv_delimiter(file_path: str) -> str:
    """
    Detect the delimiter used in a CSV. Returns ',' if unsure.
    """
    try:
        with open(file_path, 'r', encoding='utf-8', newline='') as f:
            sample = f.read(4096)
            if not sample:
                return ','
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|:")
                return dialect.delimiter or ','
            except Exception:
                counts = {sep: sample.count(sep) for sep in [';', ',', '\t', '|', ':']}
                best = max(counts.items(), key=lambda kv: kv[1])
                return best[0] if best[1] > 0 else ','
    except Exception:
        return ','


def load_and_preprocess_data(
    file_path: str,
    app_config_instance: Optional[AppConfig] = None
) -> Optional[pd.DataFrame]:
    """
    Loads data from a CSV or Excel file, standardizes column names, initializes
    new pipeline columns, and applies initial phone number normalization.

    This function supports "smart reading" for files where the exact number of
    data rows is unknown. It stops reading after encountering a configurable
    number of consecutive empty rows.

    Args:
        file_path (str): The path to the input CSV or Excel file.
        app_config_instance (Optional[AppConfig]): An optional instance of
            AppConfig. If not provided, a new one will be instantiated.
            This allows for passing a pre-configured AppConfig instance,
            useful for testing or specific runtime configurations.

    Returns:
        pd.DataFrame | None: The processed DataFrame with standardized and
        new columns. Returns None if a critical error occurs during
        loading (e.g., file not found, unsupported file type).
        Returns an empty DataFrame if the input file is empty or
        contains no valid data after applying skip/read limits.
    """
    current_config_instance: AppConfig
    if app_config_instance:
        current_config_instance = app_config_instance
    else:
        # Instantiate AppConfig if not provided; it handles .env loading.
        current_config_instance = AppConfig()

    skip_rows_val: Optional[int] = None
    nrows_val: Optional[int] = None
    # Default for smart read, will be overridden by config if available.
    consecutive_empty_rows_to_stop: int = 3

    # Load row range configurations if they exist in AppConfig
    # These attributes might not exist if AppConfig is a minimal version or old.
    if hasattr(current_config_instance, 'skip_rows_config'):
        skip_rows_val = current_config_instance.skip_rows_config
    if hasattr(current_config_instance, 'nrows_config'):
        nrows_val = current_config_instance.nrows_config
    if hasattr(current_config_instance, 'consecutive_empty_rows_to_stop'):
        consecutive_empty_rows_to_stop = current_config_instance.consecutive_empty_rows_to_stop

    log_message_parts = []
    if skip_rows_val is not None:
        # Clarify that skip_rows_val is 0-indexed from the start of the file (header is row 0)
        log_message_parts.append(f"skipping first {skip_rows_val} rows (0-indexed from file start, header is row 0)")
    if nrows_val is not None:
        log_message_parts.append(f"reading a maximum of {nrows_val} data rows (after any skipped rows)")

    # Smart read is active if nrows_val is not set (open-ended read) AND
    # consecutive_empty_rows_to_stop is a positive number.
    smart_read_active = (nrows_val is None and consecutive_empty_rows_to_stop > 0)
    if smart_read_active:
        log_message_parts.append(f"smart read active (will stop after {consecutive_empty_rows_to_stop} consecutive empty data rows)")

    if log_message_parts:
        logger.info(f"Data loading configuration: {', '.join(log_message_parts)}.")
    else:
        logger.info("No specific row range configured; attempting to load all rows (or smart read if enabled by default).")

    # pandas_skiprows_arg is for when not using smart read, or for the initial skip
    # before smart reading begins. It refers to lines in the file (1-indexed for skiprows list).
    # Pandas `skiprows` parameter:
    # - int: number of lines to skip from start of file (0 means no lines, header is line 0)
    # - list-like: 0-indexed line numbers to skip. So [0] skips header. [1] skips first data row.
    # Our `skip_rows_val` is intended to mean "number of rows to skip *after* the header".
    # So, if skip_rows_val = 1, we want pandas to skip file line 1 (the first data row).
    pandas_skiprows_arg: Union[int, List[int]]
    if skip_rows_val is not None and skip_rows_val > 0:
        # To skip `skip_rows_val` data rows, we need to skip file lines 1 to `skip_rows_val`.
        pandas_skiprows_arg = list(range(1, skip_rows_val + 1))
    else:
        pandas_skiprows_arg = 0 # Skip no lines after the header (header=0 means header is row 0)

    df: Optional[pd.DataFrame] = None

    try:
        logger.info(f"Attempting to load data from: {file_path}")

        if smart_read_active:
            logger.info(f"Smart read enabled. Max consecutive empty rows to stop: {consecutive_empty_rows_to_stop}")
            header: Optional[List[str]] = None
            data_rows: List[List[Any]] = []
            empty_row_counter = 0

            # actual_data_rows_to_skip refers to 0-indexed data rows (rows *after* the header)
            actual_data_rows_to_skip = skip_rows_val if skip_rows_val is not None else 0

            if file_path.endswith(('.xls', '.xlsx')):
                workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
                sheet = workbook.active

                if sheet is None:
                    logger.warning(f"Excel file {file_path} does not have an active sheet or is empty. Returning empty DataFrame.")
                    header = None # Ensure header is None for empty DataFrame creation
                    data_rows = [] # Ensure data_rows is empty
                else:
                    rows_iter = sheet.iter_rows()

                    # 1. Read header from the first row of the sheet
                    try:
                        header_row_values = next(rows_iter)
                        header = [str(cell.value) if cell.value is not None else '' for cell in header_row_values]
                        logger.info(f"Excel header read: {header}")
                    except StopIteration: # Handles empty sheet
                        logger.warning(f"Excel file {file_path} seems empty (no header row found).")
                        # header will remain None, data_rows empty.

                    # Only proceed to read data rows if header was successfully read
                    if header is not None:
                        # 2. Skip initial data rows (after header)
                        for _ in range(actual_data_rows_to_skip):
                            try:
                                next(rows_iter)
                            except StopIteration:
                                logger.info(f"Reached end of Excel file while skipping initial {actual_data_rows_to_skip} data rows.")
                                break

                        # 3. Read data rows with empty row detection
                        for row_idx, row_values_tuple in enumerate(rows_iter):
                            current_row_values = [cell.value for cell in row_values_tuple]
                            if _is_row_empty(current_row_values):
                                empty_row_counter += 1
                                if empty_row_counter >= consecutive_empty_rows_to_stop:
                                    logger.info(f"Stopping Excel read: Found {empty_row_counter} consecutive empty rows at data row index {actual_data_rows_to_skip + row_idx}.")
                                    break
                            else:
                                empty_row_counter = 0 # Reset counter on non-empty row
                                data_rows.append(current_row_values)

                # Create DataFrame from collected header and data_rows
                if header: # If header was read
                    df = pd.DataFrame(data_rows, columns=header)
                    logger.info(f"Smart read from Excel resulted in {len(data_rows)} data rows.")
                elif not data_rows and header is None: # Handles case where sheet was None or truly empty
                    df = pd.DataFrame() # Create an empty DataFrame
                    logger.info("Smart read from Excel: sheet was None or empty, created empty DataFrame.")
                # else: # This case (header is None but data_rows has content) should ideally not occur.
                # df = pd.DataFrame(data_rows) # Fallback if header is somehow None but data exists

            elif file_path.endswith('.csv'):
                detected_sep = _detect_csv_delimiter(file_path)
                logger.info(f"Detected CSV delimiter (smart read): '{detected_sep}' for file {file_path}")
                with open(file_path, mode='r', encoding='utf-8', newline='') as csvfile:
                    reader = csv.DictReader(csvfile, delimiter=detected_sep)

                    # 1. Read header
                    try:
                        header = list(reader.fieldnames) if reader.fieldnames else None
                        logger.info(f"CSV header read: {header}")
                    except StopIteration: # Handles empty CSV
                        logger.warning(f"CSV file {file_path} seems empty (no header row).")
                        return pd.DataFrame() # Return empty DataFrame

                    # 2. Skip initial data rows
                    for _ in range(actual_data_rows_to_skip):
                        try:
                            next(reader)
                        except StopIteration:
                            logger.info(f"Reached end of CSV file while skipping initial {actual_data_rows_to_skip} data rows.")
                            break

                    # 3. Read data with empty row detection
                    for row_idx, current_row_values in enumerate(reader):
                        # csv.reader can yield empty lists for completely blank lines
                        is_empty = not current_row_values or _is_row_empty(current_row_values.values())

                        if is_empty:
                            empty_row_counter += 1
                            if empty_row_counter >= consecutive_empty_rows_to_stop:
                                logger.info(f"Stopping CSV read: Found {empty_row_counter} consecutive empty rows at data row index {actual_data_rows_to_skip + row_idx}.")
                                break
                        else:
                            empty_row_counter = 0
                            data_rows.append(list(current_row_values.values()))

                if header: # If header was read
                    df = pd.DataFrame(data_rows, columns=header)
                # else: # Should not be reached if header read was successful and file wasn't empty
                # df = pd.DataFrame(data_rows)
                logger.info(f"Smart read from CSV resulted in {len(data_rows)} data rows.")
            else:
                logger.error(f"Unsupported file type for smart read: {file_path}. Please use CSV or Excel.")
                return None
        else: # Standard pandas read (fixed range or smart read disabled)
            logger.info(f"Using standard pandas read. Pandas skiprows argument: {pandas_skiprows_arg}, nrows: {nrows_val}")
            # keep_default_na=False and na_filter=False to prevent pandas from interpreting empty strings as NaN
            if file_path.endswith('.csv'):
                detected_sep = _detect_csv_delimiter(file_path)
                logger.info(f"Detected CSV delimiter: '{detected_sep}' for file {file_path}")
                df = pd.read_csv(file_path, header=0, skiprows=pandas_skiprows_arg, nrows=nrows_val, keep_default_na=False, na_filter=False, sep=detected_sep)
            elif file_path.endswith(('.xls', '.xlsx')):
                df = pd.read_excel(file_path, header=0, skiprows=pandas_skiprows_arg, nrows=nrows_val, keep_default_na=False, na_filter=False)
            else:
                logger.error(f"Unsupported file type: {file_path}. Please use CSV or Excel.")
                return None

        if df is None: # Should only occur if smart read was attempted for an unsupported file type
            logger.error(f"DataFrame is None after loading attempt for {file_path}. This indicates an issue with the loading logic or unsupported file for smart read.")
            return None

        # Normalize column names: strip whitespace and remove BOM if present
        try:
            cleaned_cols = []
            for c in list(df.columns):
                name = str(c) if c is not None else ''
                name = name.replace('\ufeff', '').strip()
                cleaned_cols.append(name)
            df.columns = cleaned_cols
        except Exception:
            pass

        logger.info(f"Columns loaded (normalized): {df.columns.tolist() if df is not None and not df.empty else 'N/A (DataFrame is None or empty)'}")

        if df.empty:
            logger.warning(f"Loaded DataFrame from {file_path} is empty. This could be due to an empty input file, all rows being skipped, or smart read stopping early.")
            # If df is empty, we still want to ensure essential columns are present for later stages.
            # The new_columns loop later will add them if they don't exist.

        # --- Post-loading processing: Apply input profile for column renaming and add new pipeline columns ---
        active_profile_name = current_config_instance.input_file_profile_name
        profile_mappings = current_config_instance.INPUT_COLUMN_PROFILES.get(active_profile_name)

        if not profile_mappings:
            logger.error(f"Input profile '{active_profile_name}' not found in AppConfig.INPUT_COLUMN_PROFILES. Falling back to 'default' profile.")
            active_profile_name = "default" # Attempt to use a default profile
            profile_mappings = current_config_instance.INPUT_COLUMN_PROFILES.get("default")
            if not profile_mappings: # This should ideally not happen if "default" is always defined in AppConfig
                 logger.error("Critical: Default input profile ('default') not found in AppConfig. Cannot map columns.")
                 return pd.DataFrame() # Return empty DataFrame as a fallback

        # Create rename map only for columns present in the DataFrame
        # Also consider case-insensitive matches for source columns to be more forgiving
        lower_to_actual = {str(c).lower(): c for c in df.columns}
        actual_rename_map = {}
        for src_name, dst_name in profile_mappings.items():
            if src_name.startswith('_'):
                continue
            if src_name in df.columns:
                actual_rename_map[src_name] = dst_name
            else:
                cand = lower_to_actual.get(str(src_name).lower())
                if cand is not None:
                    actual_rename_map[cand] = dst_name

        if actual_rename_map:
             df.rename(columns=actual_rename_map, inplace=True)
        # If the input already contains canonical columns, profile renaming can create duplicates.
        # Coalesce duplicates now so downstream row access always yields scalars, not Series.
        df = _coalesce_duplicate_columns(df)
        logger.info(f"DataFrame columns after renaming (using profile: '{active_profile_name}'): {df.columns.tolist()}")

        # Drop accidental header rows in data (common in scraped/merged lists)
        try:
            header_like_company = {"Aussteller", "Firma"}
            header_like_url = {"Website", "WebsiteUnternehmen", "Website Unternehmen"}
            if 'CompanyName' in df.columns:
                df = df[~df['CompanyName'].astype(str).str.strip().isin(header_like_company)]
            if 'GivenURL' in df.columns:
                df = df[~df['GivenURL'].astype(str).str.strip().isin(header_like_url)]
        except Exception:
            pass

        # Define and initialize new columns required by the pipeline
        new_columns = [
            "NormalizedGivenPhoneNumber", "ScrapingStatus",
            "Overall_VerificationStatus", "Original_Number_Status",
            "Primary_Number_1", "Primary_Type_1", "Primary_SourceURL_1",
            "Secondary_Number_1", "Secondary_Type_1", "Secondary_SourceURL_1",
            "Secondary_Number_2", "Secondary_Type_2", "Secondary_SourceURL_2",
            "RunID", "TargetCountryCodes", "is_b2b", "serves_1000"
        ]

        current_run_id = str(uuid.uuid4()) # Generate a unique RunID for this processing batch

        for col in new_columns:
            if col not in df.columns:
                if col == "RunID":
                    df[col] = current_run_id
                elif col == "TargetCountryCodes":
                    # Initialize with default target countries; robust for empty df
                    df[col] = pd.Series([["DE", "AT", "CH"] for _ in range(len(df))] if not df.empty else [], dtype=object)
                elif col in ["ScrapingStatus", "Overall_VerificationStatus", "Original_Number_Status"]:
                    df[col] = "Pending" # Default status
                elif col.startswith("Primary_") or col.startswith("Secondary_"):
                    df[col] = None # Initialize phone/type/source columns as None
                else:
                    df[col] = None # Default for other new columns
        
        logger.info(f"Successfully loaded and structured data from {file_path}. DataFrame shape: {df.shape}")

        return df
    except FileNotFoundError:
        logger.error(f"Error: The file {file_path} was not found.")
        return None
    except pd.errors.EmptyDataError: # This might be caught by smart read logic earlier for empty files
        logger.error(f"Error: The file {file_path} is empty (pandas EmptyDataError).")
        # Return an empty DataFrame and None for the phone column name, consistent with other error paths.
        return pd.DataFrame()
    except Exception as e:
        logger.error(f"An unexpected error occurred while loading data from {file_path}: {e}", exc_info=True)
        return None